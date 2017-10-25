"""
PART OF THE QUANTITATIVE BEHAVIORAL ANALYSIS AND MODELING (QBAM) PROJECT

Haozhe Shan
Mason Laboratory, University of Chicago
September 2016

=============================================================================
This is the FILE PROCESSOR V5.

It converts output files from Ethovision (Excel spreadsheets) into Python pickle files that contain custom-built objects named "processed_data". These objects contain
all relevant information regarding the original raw data.
=============================================================================
"""
from BLC_util import processed_data, loadseries
import numpy as np
import time
from openpyxl import load_workbook
import os.path
import cPickle as pickle
#  =================================
#  BASICS OF THE EXCEL FILES
#  Open any Excel file in the group to label the follows.
#  =================================
time_column = 'A'
X_column = 'C'
Y_column = 'D'
area_column = 'E'
velocity_column = 'I'
ConditionID_cell = 'B33'
Trapped_ID_cell = 'B5'
Date_cell = 'B27'
opener_cell = 'B35'
Group_ID = 'B32'
rat_count = 32
day_count = 5

#  =================================
#  INDEX OF DATA FILES
#  Which files do you need to import? Get their indexes from the file names.
#  =================================
loading_start=2
loading_end=160

'''
=================
DEFINITIONS
=================
'''

group_name = secondary_analysis_target_header


#%%


starttime = time.time()
for z in range(loading_start, loading_end+1):

    loading_video_number = z

    if z < 10:
        name = group_name + '  %s.xlsx' % loading_video_number
    elif 9 < z < 100:
        name = group_name + ' %s.xlsx' % loading_video_number
    else:
        name = group_name + '%s.xlsx' % loading_video_number

    if os.path.isfile(name) is False:
        print('File not found')
        continue
    print "Currently processing: "+name
    book_in_processing = load_workbook(name)

    for m in xrange(number_of_rats_per_video):
        sheet_in_processing=book_in_processing.worksheets[m]
        if z == loading_start and m == 0:
            first_row = int(sheet_in_processing['B1'].value)+1
        print sheet_in_processing.title
        last_row = sheet_in_processing.max_row


        condition = sheet_in_processing[ConditionID_cell].value
        trapped_id = sheet_in_processing[Trapped_ID_cell].value
        date = sheet_in_processing[Date_cell].value
        opener = sheet_in_processing[opener_cell].value
        group = sheet_in_processing[Group_ID].value

        temporary_loaded_vector, length = loadseries(first_row,last_row,time_column,sheet_in_processing)
        if m==0:
            timeline = np.zeros((length,number_of_rats_per_video))
            Xseries = timeline.copy()
            Yseries = timeline.copy()
            area = timeline.copy()
            velocity = timeline.copy()


        timeline[:,m] = temporary_loaded_vector
        temporary_loaded_vector,length=loadseries(first_row,last_row,X_column,sheet_in_processing)
        Xseries[:,m] = temporary_loaded_vector
        temporary_loaded_vector,length=loadseries(first_row,last_row,Y_column,sheet_in_processing)
        Yseries[:,m] = temporary_loaded_vector
        temporary_loaded_vector,length=loadseries(first_row,last_row,area_column,sheet_in_processing)
        area[:,m] = temporary_loaded_vector
        temporary_loaded_vector,length=loadseries(first_row,last_row,velocity_column,sheet_in_processing)
        velocity[:,m] = temporary_loaded_vector

    loaded_data_object = processed_data(name, condition, group, trapped_id, date, timeline[:, 0], Xseries[:, 0],
                                      Yseries[:, 0], velocity[:, 0], opener)

    loaded_data_filename = group_name + "%s.pkl" % loading_video_number
    print "Time elapsed: %05s sec" % (time.time()-starttime)
    pickle.dump(loaded_data_object, open(loaded_data_filename, 'wb'))

print 'File Processing Complete.'

#  %% File treater: lowpass filtering and smoothing


read_pickle_file_header = "Raw data-Group D, D1-12 Uninjected-Trial    "
save_treated_file_header = "SMOOTHFILTER_Raw data-Group D, D1-12 Uninjected-Trial    "
sampling_freq=15
lowpass_freq=0
highpass_freq=3
order=2

start_treating_at=1
end_treating_at=80

b, a = butter_bandpass(lowpass_freq, highpass_freq, sampling_freq, order=order)
treating_start_time=time.time()

for i in range(start_treating_at,end_treating_at+1):
    analyze_video_number=i
    picklename=read_pickle_file_header+'%s.pkl' % analyze_video_number

    if os.path.isfile(picklename) == False:
        continue
    print "Currently processing "+picklename

    data_to_be_treated=pickle.load(open(picklename,'rb'))
    pretreated_x_series=data_to_be_treated.x0
    pretreated_y_series=data_to_be_treated.y0
    pretreated_velocity_series=data_to_be_treated.v0
    smooth_x_series=scipy.ndimage.gaussian_filter1d(pretreated_x_series,15)
    smooth_y_series=scipy.ndimage.gaussian_filter1d(pretreated_y_series,15)
    smooth_velocity_series=scipy.ndimage.gaussian_filter1d(pretreated_velocity_series,15)

    filtered_x_series = butter_bandpass_filter(smooth_x_series, lowpass_freq, highpass_freq, sampling_freq, order=order)
    filtered_y_series = butter_bandpass_filter(smooth_y_series, lowpass_freq, highpass_freq, sampling_freq, order=order)
    filtered_velocity_series = butter_bandpass_filter(smooth_velocity_series, lowpass_freq, highpass_freq, sampling_freq, order=order)

    data_to_be_treated.x0=filtered_x_series
    data_to_be_treated.y0=filtered_y_series
    data_to_be_treated.v0=filtered_velocity_series

    print "Time elapsed: %05s sec" % (time.time()-treating_start_time)
    picklename=save_treated_file_header+"%s.pkl" % analyze_video_number
    pickle.dump(data_to_be_treated,open(picklename,'wb'))

